from openpyxl import Workbook, load_workbook
import os

file_name = "output.xlsx"

# Check if file exists
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws.append(["Name", "Age", "Role"])  # Header only once

# Take inputs
name = input("Enter name: ")
age = input("Enter age: ")
role = input("Enter role: ")

# Append ONE ROW of data
ws.append([name, age, role])

# Save file
wb.save(file_name)
print("Data added successfully!")
