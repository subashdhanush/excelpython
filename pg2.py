from openpyxl import Workbook, load_workbook
import os

file_name = "output.xlsx"

# Check if file already exists
if os.path.exists(file_name):
    wb = load_workbook(file_name)   # Open existing file
    ws = wb.active
else:
    wb = Workbook()                 # Create new file
    ws = wb.active
    ws.title = "Inputs"
    ws.append(["Field", "Value"])   # Header only once

# Take inputs
name = input("Enter name: ")
age = input("Enter age: ")
role = input("Enter role: ")

# Append new data
ws.append(["Name", name])
ws.append(["Age", age])
ws.append(["Role", role])

# Save file
wb.save(file_name)
print("Data added successfully!")
